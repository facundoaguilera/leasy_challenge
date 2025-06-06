from django.db.models import OuterRef, Subquery, Q, Prefetch
from contracts.models import Contract
from clients.models import Client
from vehicles.models import Vehicle
import csv
from openpyxl import load_workbook
from django.utils.dateparse import parse_date
import unicodedata
from datetime import datetime, date

REQUIRED_COLUMNS = [
    "nombres",
    "apellidos",
    "numero de documento",
    "inicio de contrato",
    "cuota semanal",
    "marca del auto",
    "modelo del auto",
    "placa del auto"
]
MAX_CHUNK_SIZE = 5 * 1024 * 1024  # 5 MB
def normalize(text):
    if not text:
        return ""
    return ''.join(
        c for c in unicodedata.normalize('NFD', text.lower().strip())
        if unicodedata.category(c) != 'Mn'
    )
def process_rows(rows, normalized_map):
    nuevos, duplicados = 0, 0
    
    for row in rows:
        doc =row.get("numero de documento")
    
        if not doc:
            continue
             
        client, created = Client.objects.get_or_create(
            document_number=doc,
            first_name =  row.get("nombres"),
            last_name = row.get("apellidos")
            
        )
        nuevos += int(created)
        duplicados += int(not created)      
        plate = row.get("placa del auto")
        vehicle, _ = Vehicle.objects.get_or_create(
            plate=plate,
            brand= row.get("marca del auto"),
            model= row.get("modelo del auto"),
            vin = row.get("vin") or None,
          
        ) 
                
        fecha_inicio = row.get("inicio de contrato")#parse_fecha(get_col("inicio de contrato"))#parse_date(get_col("inicio de contrato") )
        
        if not Contract.objects.filter(client=client).exists() and fecha_inicio:
            ciclo = "weekly"
            for col in normalized_map:
                if "cuota" in col:
                    if "semanal" in col:
                        ciclo = "weekly"
                    elif "quincenal" in col:
                        ciclo = "biweekly"
                    elif "mensual" in col:
                        ciclo = "monthly"
                    break       
            monto = (
                row.get("cuota semanal") or
                row.get("cuota quincenal") or
                row.get("cuota mensual")
            )       
            Contract.objects.create(
                client=client,
                vehicle=vehicle,
                billing_cycle=ciclo,
                amount=monto,
                start_date=fecha_inicio,
                active=True
            )  
                          
    return nuevos, duplicados
        

class DashboardService:
    @staticmethod
    def get_filtered_clients(query=None):
        latest_contract = Contract.objects.filter(
            client=OuterRef("pk")
        ).order_by('-active', '-start_date')

        clients = Client.objects.annotate(
            latest_contract_id=Subquery(latest_contract.values("id")[:1])
        ).prefetch_related(
            Prefetch("contracts", queryset=Contract.objects.select_related("vehicle"), to_attr="all_contracts")
        ).order_by("last_name", "first_name")

        if query:
            clients = clients.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(document_number__icontains=query) |
                Q(contracts__vehicle__brand__icontains=query) |
                Q(contracts__vehicle__model__icontains=query) |
                Q(contracts__vehicle__plate__icontains=query)
            ).distinct()
        
        return clients
    
   
    @staticmethod
    def process_file(file):
        try:
            if file.name.endswith(".csv"):
                nuevos, duplicados = 0, 0
                position = 0
                while True:
                    file.seek(position)
                    chunk = file.read(MAX_CHUNK_SIZE)
                    if not chunk:
                        break
                    line = file.readline()
                    chunk += line
                    position = file.tell()

                    decoded = chunk.decode("utf-8").splitlines()
                    reader = csv.DictReader(decoded)
                    raw_headers = reader.fieldnames
                    rows_csv = list(reader)
                    
                    normalized_map = [normalize(h) for h in raw_headers if h]
                    missing = [col for col in REQUIRED_COLUMNS if col not in normalized_map]
                    
                    if missing:
                        
                        return {"error": f"Faltan columnas requeridas: {', '.join(missing)}"}
                    rows=[]
                    for row in rows_csv:
                        row_dict = dict(zip(normalized_map,row.values()))
                        rows.append(row_dict)
                    
                    n, d = process_rows(rows, normalized_map)
                    nuevos += n
                    duplicados += d

                return nuevos, duplicados

            elif file.name.endswith(".xlsx"):
                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active
                first_row = next(ws.iter_rows(min_row=1, max_row=1))
                raw_headers = [str(cell.value).strip().lower() for cell in first_row]
                normalized_map = [normalize(h) for h in raw_headers if h ]
                missing = [col for col in REQUIRED_COLUMNS if col not in normalized_map]
                if missing:
                    return {"error": f"Faltan columnas requeridas: {', '.join(missing)}"}

                total_rows = ws.max_row - 1
                estimated_row_size = file.size / total_rows if total_rows else 100
                rows_per_batch = int(MAX_CHUNK_SIZE / estimated_row_size) or 1

                batch = []
                nuevos, duplicados = 0, 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    row_dict = dict(zip(normalized_map, row))
                    batch.append(row_dict)
                    if len(batch) >= rows_per_batch:
                        n, d = process_rows(batch, normalized_map)
                        nuevos += n
                        duplicados += d
                        batch = []

                if batch:
                    n, d = process_rows(batch, normalized_map)
                    nuevos += n
                    duplicados += d

                return nuevos, duplicados

            else:
                return {"error": "Formato no soportado. Usa .csv o .xlsx"}

        except Exception as e:
            return {"error": f"Ocurrió un error procesando el archivo: {str(e)}"}
